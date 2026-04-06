"""Ad Campaign Generator — 7-step pipeline for Yandex Direct campaigns."""

from app.config import USD_TO_RUB
import io
import json
import logging
import os
from datetime import datetime

import httpx
from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db, async_session
from app.middleware.auth import get_current_user
from app.models.campaign import Campaign
from app.models.user import User
from app.services.ai_router import get_openrouter_model

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/campaigns", tags=["campaigns"])

COST_USD = 0.31  # ~29 RUB


class CreateCampaignRequest(BaseModel):
    niche: str = Field(max_length=500)
    url: str | None = Field(default=None, max_length=500)
    budget_rub: int = Field(default=30000, ge=5000, le=1000000)
    region: str = Field(default="Россия", max_length=100)


STEPS = [
    "Анализ ниши и конкурентов",
    "Генерация ключевых слов",
    "Кластеризация и группировка",
    "Генерация минус-слов",
    "Создание объявлений",
    "Оптимизация и рекомендации",
    "Финализация кампании",
]


async def _ai_call(prompt: str, system: str = "", model: str = "gpt-4o-mini") -> str:
    """Make AI call via OpenRouter."""
    openrouter_key = os.getenv("OPENROUTER_API_KEY")
    if not openrouter_key:
        raise Exception("OPENROUTER_API_KEY not set")

    messages = []
    if system:
        messages.append({"role": "system", "content": system})
    messages.append({"role": "user", "content": prompt})

    async with httpx.AsyncClient(timeout=60) as client:
        resp = await client.post(
            "https://openrouter.ai/api/v1/chat/completions",
            headers={"Authorization": f"Bearer {openrouter_key}"},
            json={"model": get_openrouter_model(model), "messages": messages, "max_tokens": 4000, "temperature": 0.3},
        )
        if resp.status_code != 200:
            raise Exception(f"AI API error: {resp.status_code}")
        return resp.json()["choices"][0]["message"]["content"]


async def run_campaign_pipeline(campaign_id: int, niche: str, url: str | None, budget: int, region: str, user_tg_id: int):
    """Background: run 7-step campaign generation."""
    result = {
        "niche_analysis": {},
        "keywords": [],
        "keyword_groups": [],
        "negative_keywords": [],
        "ads": [],
        "recommendations": [],
        "summary": {},
    }

    try:
        # Pre-check balance
        async with async_session() as db:
            u = await db.execute(select(User).where(User.telegram_id == user_tg_id))
            user = u.scalar_one_or_none()
            if user and float(user.balance_usd or 0) < COST_USD:
                r = await db.execute(select(Campaign).where(Campaign.id == campaign_id))
                c = r.scalar_one_or_none()
                if c: c.status = "failed"; c.result = {"error": "Недостаточно средств"}
                await db.commit()
                return

        # Step 1: Niche analysis
        async with async_session() as db:
            r = await db.execute(select(Campaign).where(Campaign.id == campaign_id))
            c = r.scalar_one_or_none()
            if c: c.step = 1; c.status = "running"; await db.commit()

        site_context = ""
        if url:
            try:
                async with httpx.AsyncClient(timeout=15) as client:
                    resp = await client.get(url, follow_redirects=True)
                    if resp.status_code == 200:
                        from html.parser import HTMLParser
                        class TextExtractor(HTMLParser):
                            def __init__(self):
                                super().__init__()
                                self.text = []
                            def handle_data(self, data):
                                self.text.append(data.strip())
                        parser = TextExtractor()
                        parser.feed(resp.text[:20000])
                        site_context = " ".join(parser.text)[:3000]
            except Exception:
                pass

        analysis = await _ai_call(
            f"Проанализируй нишу для рекламной кампании в Яндекс Директ.\n\nНиша: {niche}\n{f'Сайт: {url}' if url else ''}\n{f'Контент сайта: {site_context[:1000]}' if site_context else ''}\nРегион: {region}\nБюджет: {budget}₽/мес\n\nОтвети JSON: {{\"target_audience\": \"описание ЦА\", \"usp\": \"УТП\", \"competitors\": [\"конкурент1\", \"конкурент2\"], \"tone\": \"тон рекламы\", \"pain_points\": [\"боль1\", \"боль2\"]}}",
            system="Ты эксперт по контекстной рекламе Яндекс Директ. Отвечай только JSON."
        )
        try:
            result["niche_analysis"] = json.loads(analysis.strip().strip("```json").strip("```"))
        except Exception as parse_err:
            logger.warning(f"JSON parse fallback: {parse_err}")
            result["niche_analysis"] = {"raw": analysis}

        # Step 2: Keywords
        async with async_session() as db:
            r = await db.execute(select(Campaign).where(Campaign.id == campaign_id))
            c = r.scalar_one_or_none()
            if c: c.step = 2; c.result = result; await db.commit()

        keywords_raw = await _ai_call(
            f"""Сгенерируй 50-80 ключевых слов для рекламы в Яндекс Директ.

Ниша: {niche}
ЦА: {result['niche_analysis'].get('target_audience', niche)}
Регион: {region}

Для каждого ключа оцени:
- wordstat_frequency: примерная месячная частотность в Яндекс Wordstat (число показов)
- estimated_cpc: примерная цена клика в рублях
- competition: low/medium/high
- intent: commercial/informational/navigational

Ответи JSON массивом: [{{"keyword": "ключ", "wordstat_frequency": 1500, "estimated_cpc": 30, "competition": "medium", "intent": "commercial"}}]""",
            system="Ты SEO/PPC специалист с опытом работы с Яндекс Wordstat и Директ. Оценивай частотность реалистично для русскоязычного рынка. Отвечай только JSON."
        )
        try:
            result["keywords"] = json.loads(keywords_raw.strip().strip("```json").strip("```"))
        except Exception as parse_err:
            logger.warning(f"JSON parse fallback: {parse_err}")
            result["keywords"] = [{"keyword": k.strip(), "intent": "commercial", "estimated_cpc": 30, "wordstat_frequency": 0, "competition": "medium"} for k in keywords_raw.split("\n") if k.strip()]

        # Step 3: Clustering
        async with async_session() as db:
            r = await db.execute(select(Campaign).where(Campaign.id == campaign_id))
            c = r.scalar_one_or_none()
            if c: c.step = 3; c.result = result; await db.commit()

        kw_list = [k.get("keyword", k) if isinstance(k, dict) else k for k in result["keywords"]][:80]
        groups_raw = await _ai_call(
            f"Сгруппируй ключевые слова в 5-10 рекламных групп для Яндекс Директ.\n\nКлючевые слова:\n{chr(10).join(kw_list)}\n\nОтвети JSON: [{{\"group_name\": \"название группы\", \"keywords\": [\"ключ1\", \"ключ2\"]}}]",
            system="Группируй по смыслу и намерению. Отвечай только JSON."
        )
        try:
            result["keyword_groups"] = json.loads(groups_raw.strip().strip("```json").strip("```"))
        except Exception as parse_err:
            logger.warning(f"JSON parse fallback: {parse_err}")
            result["keyword_groups"] = [{"group_name": "Основная", "keywords": kw_list}]

        # Step 4: Negative keywords
        async with async_session() as db:
            r = await db.execute(select(Campaign).where(Campaign.id == campaign_id))
            c = r.scalar_one_or_none()
            if c: c.step = 4; c.result = result; await db.commit()

        neg_raw = await _ai_call(
            f"Сгенерируй 100-200 минус-слов для рекламной кампании Яндекс Директ.\n\nНиша: {niche}\nКлючевые слова: {', '.join(kw_list[:30])}\n\nОтвети JSON массивом строк: [\"минус1\", \"минус2\", ...]",
            system="Генерируй минус-слова чтобы отсечь нерелевантный трафик. Отвечай только JSON."
        )
        try:
            result["negative_keywords"] = json.loads(neg_raw.strip().strip("```json").strip("```"))
        except Exception as parse_err:
            logger.warning(f"JSON parse fallback: {parse_err}")
            result["negative_keywords"] = [w.strip().strip('"').strip("- ") for w in neg_raw.split("\n") if w.strip()]

        # Step 5: Ads
        async with async_session() as db:
            r = await db.execute(select(Campaign).where(Campaign.id == campaign_id))
            c = r.scalar_one_or_none()
            if c: c.step = 5; c.result = result; await db.commit()

        ads = []
        for group in result["keyword_groups"][:8]:
            ad_raw = await _ai_call(
                f"Создай рекламное объявление для Яндекс Директ.\n\nГруппа: {group['group_name']}\nКлючевые слова: {', '.join(group.get('keywords', [])[:10])}\nНиша: {niche}\nУТП: {result['niche_analysis'].get('usp', '')}\n\nОтвети JSON: {{\"title1\": \"заголовок до 35 символов\", \"title2\": \"доп заголовок до 30 символов\", \"text\": \"текст объявления до 81 символа\", \"sitelinks\": [{{\"title\": \"ссылка\", \"url\": \"/page\"}}]}}",
                system="Ты копирайтер Яндекс Директ. Соблюдай лимиты символов. Отвечай только JSON."
            )
            try:
                ad = json.loads(ad_raw.strip().strip("```json").strip("```"))
                ad["group"] = group["group_name"]
                ads.append(ad)
            except Exception as parse_err:
                logger.warning(f"JSON parse fallback: {parse_err}")
                ads.append({"group": group["group_name"], "title1": niche[:35], "title2": "Закажите сейчас", "text": f"{niche}. Доставка. Гарантия.", "raw": ad_raw})
        result["ads"] = ads

        # Step 6: Recommendations
        async with async_session() as db:
            r = await db.execute(select(Campaign).where(Campaign.id == campaign_id))
            c = r.scalar_one_or_none()
            if c: c.step = 6; c.result = result; await db.commit()

        recs_raw = await _ai_call(
            f"Дай 5-7 рекомендаций по оптимизации рекламной кампании.\n\nНиша: {niche}\nБюджет: {budget}₽/мес\nКлючей: {len(kw_list)}\nГрупп: {len(result['keyword_groups'])}\nМинус-слов: {len(result['negative_keywords'])}\n\nОтвети JSON: [{{\"title\": \"рекомендация\", \"description\": \"описание\", \"priority\": \"high|medium|low\"}}]",
            system="Ты PPC-эксперт. Отвечай только JSON."
        )
        try:
            result["recommendations"] = json.loads(recs_raw.strip().strip("```json").strip("```"))
        except Exception as parse_err:
            logger.warning(f"JSON parse fallback: {parse_err}")
            result["recommendations"] = [{"title": "Оптимизация", "description": recs_raw, "priority": "medium"}]

        # Step 7: Summary
        result["summary"] = {
            "total_keywords": len(kw_list),
            "total_groups": len(result["keyword_groups"]),
            "total_negative": len(result["negative_keywords"]),
            "total_ads": len(result["ads"]),
            "estimated_budget": budget,
            "region": region,
        }

        # Final save
        async with async_session() as db:
            r = await db.execute(select(Campaign).where(Campaign.id == campaign_id))
            c = r.scalar_one_or_none()
            if c:
                c.step = 7
                c.status = "completed"
                c.result = result
                c.total_keywords = len(kw_list)
                c.total_ads = len(ads)
                c.cost_usd = COST_USD
                c.completed_at = datetime.utcnow()
            # Deduct cost
            u = await db.execute(select(User).where(User.telegram_id == user_tg_id))
            user = u.scalar_one_or_none()
            if user:
                user.balance_usd = max(0, float(user.balance_usd or 0) - COST_USD)
            await db.commit()

    except Exception as e:
        logger.error(f"Campaign pipeline error: {e}")
        async with async_session() as db:
            r = await db.execute(select(Campaign).where(Campaign.id == campaign_id))
            c = r.scalar_one_or_none()
            if c:
                c.status = "failed"
                c.result = {**(c.result or {}), "error": str(e)[:500]}
            await db.commit()


@router.post("/create")
async def create_campaign(
    body: CreateCampaignRequest,
    background_tasks: BackgroundTasks,
    tg_user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Start campaign generation."""
    tg_id = tg_user["id"]

    # Check balance
    user = await db.execute(select(User).where(User.telegram_id == tg_id))
    u = user.scalar_one_or_none()
    if u and float(u.balance_usd or 0) < COST_USD:
        raise HTTPException(402, f"Недостаточно средств. Нужно ~{int(COST_USD * USD_TO_RUB)}₽")

    # Limit active campaigns
    active = await db.scalar(
        select(func.count()).select_from(Campaign).where(Campaign.user_tg_id == tg_id, Campaign.status == "running")
    )
    if active >= 2:
        raise HTTPException(429, "Максимум 2 одновременных кампании")

    campaign = Campaign(
        user_tg_id=tg_id,
        name=f"Кампания: {body.niche[:100]}",
        niche=body.niche,
        url=body.url,
        status="running",
    )
    db.add(campaign)
    await db.flush()

    background_tasks.add_task(run_campaign_pipeline, campaign.id, body.niche, body.url, body.budget_rub, body.region, tg_id)

    return {"id": campaign.id, "status": "running"}


@router.get("/list")
async def list_campaigns(
    tg_user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """List user's campaigns."""
    result = await db.execute(
        select(Campaign).where(Campaign.user_tg_id == tg_user["id"]).order_by(Campaign.created_at.desc()).limit(20)
    )
    campaigns = result.scalars().all()
    return {
        "campaigns": [
            {
                "id": c.id, "name": c.name, "status": c.status, "step": c.step,
                "total_keywords": c.total_keywords, "total_ads": c.total_ads,
                "cost_rub": round(c.cost_usd * USD_TO_RUB),
                "created_at": c.created_at.isoformat() if c.created_at else None,
            }
            for c in campaigns
        ]
    }


IMPROVE_PRESETS = {
    "ctr": {"label": "📈 Повысить CTR", "prompt": "Перепиши объявление с интригой и срочностью для максимального CTR. Добавь цифры и ограничение по времени."},
    "cta": {"label": "🎯 Усилить CTA", "prompt": "Усиль призыв к действию. Сделай его конкретным и побуждающим немедленно кликнуть."},
    "offer": {"label": "💪 Усилить оффер", "prompt": "Усиль выгоду и предложение. Добавь бонусы, гарантии, скидки."},
    "numbers": {"label": "🔢 Добавить цифры", "prompt": "Добавь конкретные цифры, факты, статистику. Замени абстрактные слова на конкретные."},
    "emotional": {"label": "❤️ Эмоциональный", "prompt": "Перепиши в эмоциональном стиле. Добавь теплоту, чувства, личное обращение."},
    "formal": {"label": "👔 Деловой тон", "prompt": "Перепиши в строго деловом стиле. Серьёзность, доверие, экспертность."},
    "keywords": {"label": "🔑 Вставить ключевики", "prompt": "Вставь ключевые слова группы в заголовок и текст для повышения релевантности."},
    "shorten": {"label": "✂️ Сократить", "prompt": "Сократи текст до минимума. Убери всё лишнее, оставь только суть."},
}

IMPROVE_COST_USD = 0.032  # ~3 RUB


class ImproveAdRequest(BaseModel):
    campaign_id: int
    ad_index: int
    preset: str


@router.post("/improve-ad")
async def improve_ad(
    body: ImproveAdRequest,
    tg_user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Improve a single ad using AI preset."""
    tg_id = tg_user["id"]
    preset = IMPROVE_PRESETS.get(body.preset)
    if not preset:
        raise HTTPException(400, "Неизвестный пресет")

    # Get campaign
    result = await db.execute(
        select(Campaign).where(Campaign.id == body.campaign_id, Campaign.user_tg_id == tg_id)
    )
    c = result.scalar_one_or_none()
    if not c or not c.result or not c.result.get("ads"):
        raise HTTPException(404, "Кампания не найдена")

    ads = c.result["ads"]
    if body.ad_index < 0 or body.ad_index >= len(ads):
        raise HTTPException(400, "Неверный индекс объявления")

    # Check balance
    user_result = await db.execute(select(User).where(User.telegram_id == tg_id))
    user = user_result.scalar_one_or_none()
    if user and float(user.balance_usd or 0) < IMPROVE_COST_USD:
        raise HTTPException(402, "Недостаточно средств (~3₽)")

    ad = ads[body.ad_index]
    original = {"title1": ad.get("title1", ""), "title2": ad.get("title2", ""), "text": ad.get("text", "")}

    try:
        improved_raw = await _ai_call(
            f"""{preset['prompt']}

Текущее объявление:
Заголовок 1: {ad.get('title1', '')}
Заголовок 2: {ad.get('title2', '')}
Текст: {ad.get('text', '')}
Группа: {ad.get('group', '')}

ВАЖНО: Заголовок 1 — максимум 35 символов. Заголовок 2 — максимум 30 символов. Текст — максимум 81 символ.

Ответь JSON: {{"title1": "...", "title2": "...", "text": "..."}}""",
            system="Ты копирайтер Яндекс Директ. Строго соблюдай лимиты символов. Отвечай только JSON."
        )
        improved = json.loads(improved_raw.strip().strip("```json").strip("```"))

        # Update ad in campaign
        ads[body.ad_index] = {**ad, **improved}
        c.result = {**c.result, "ads": ads}

        # Deduct cost
        if user:
            user.balance_usd = max(0, float(user.balance_usd or 0) - IMPROVE_COST_USD)
        c.cost_usd = (c.cost_usd or 0) + IMPROVE_COST_USD

        return {
            "ok": True,
            "original": original,
            "improved": improved,
            "cost_rub": round(IMPROVE_COST_USD * USD_TO_RUB),
        }
    except Exception as e:
        logger.error(f"Ad improve error: {e}")
        raise HTTPException(502, "Ошибка улучшения")


@router.get("/{campaign_id}")
async def get_campaign(
    campaign_id: int,
    tg_user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Get campaign details with full result."""
    result = await db.execute(
        select(Campaign).where(Campaign.id == campaign_id, Campaign.user_tg_id == tg_user["id"])
    )
    c = result.scalar_one_or_none()
    if not c:
        raise HTTPException(404, "Кампания не найдена")

    return {
        "id": c.id, "name": c.name, "niche": c.niche, "url": c.url,
        "status": c.status, "step": c.step,
        "result": c.result,
        "total_keywords": c.total_keywords, "total_ads": c.total_ads,
        "cost_rub": round(c.cost_usd * USD_TO_RUB),
        "created_at": c.created_at.isoformat() if c.created_at else None,
        "completed_at": c.completed_at.isoformat() if c.completed_at else None,
    }


@router.get("/{campaign_id}/export")
async def export_campaign_xlsx(
    campaign_id: int,
    tg_user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Export campaign to XLSX for Yandex Direct import."""
    result = await db.execute(
        select(Campaign).where(Campaign.id == campaign_id, Campaign.user_tg_id == tg_user["id"])
    )
    c = result.scalar_one_or_none()
    if not c or not c.result:
        raise HTTPException(404, "Кампания не найдена")

    try:
        from openpyxl import Workbook
        wb = Workbook()

        # Sheet 1: Keywords
        ws_kw = wb.active
        ws_kw.title = "Ключевые слова"
        ws_kw.append(["Группа", "Ключевое слово", "Частотность (Wordstat)", "CPC (₽)", "Конкуренция", "Intent"])
        # Map keywords to their data
        kw_data = {k.get("keyword", ""): k for k in c.result.get("keywords", []) if isinstance(k, dict)}
        for group in c.result.get("keyword_groups", []):
            for kw in group.get("keywords", []):
                data = kw_data.get(kw, {})
                ws_kw.append([group["group_name"], kw, data.get("wordstat_frequency", ""), data.get("estimated_cpc", ""), data.get("competition", ""), data.get("intent", "")])

        # Sheet 2: Negative keywords
        ws_neg = wb.create_sheet("Минус-слова")
        ws_neg.append(["Минус-слово"])
        for neg in c.result.get("negative_keywords", []):
            ws_neg.append([neg])

        # Sheet 3: Ads
        ws_ads = wb.create_sheet("Объявления")
        ws_ads.append(["Группа", "Заголовок 1", "Заголовок 2", "Текст"])
        for ad in c.result.get("ads", []):
            ws_ads.append([ad.get("group", ""), ad.get("title1", ""), ad.get("title2", ""), ad.get("text", "")])

        # Sheet 4: Recommendations
        ws_rec = wb.create_sheet("Рекомендации")
        ws_rec.append(["Приоритет", "Рекомендация", "Описание"])
        for rec in c.result.get("recommendations", []):
            ws_rec.append([rec.get("priority", ""), rec.get("title", ""), rec.get("description", "")])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=campaign-{campaign_id}.xlsx"},
        )
    except ImportError:
        raise HTTPException(500, "openpyxl не установлен")
